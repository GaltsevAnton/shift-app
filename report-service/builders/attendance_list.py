"""
打刻一覧（フラットリスト） — одна строка на каждую пробивку.
Колонки: 職種・役職 | 部署 | 氏名 | 日付 | 曜日 | 種別 | 時刻 | シフト予定
Сортировка: по дате, затем по氏名, затем по времени пробивки.
"""

import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from models import AttendanceReportRequest

WD_JA     = ["日", "月", "火", "水", "木", "金", "土"]
FONT_NAME = "メイリオ"

C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_ROW_EVEN  = "F2F2F2"
C_GRID      = "AAAAAA"

RECORD_LABELS = {
    "CLOCK_IN":    "出勤",
    "CLOCK_OUT":   "退勤",
    "BREAK_START": "休憩開始",
    "BREAK_END":   "休憩終了",
}


def _thin(color=C_GRID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)

def _font(bold=False, color="000000", size=9, name=FONT_NAME):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _parse_dt(iso_dt: str | None):
    if not iso_dt:
        return None
    try:
        return datetime.fromisoformat(iso_dt)
    except Exception:
        return None


def build(req: AttendanceReportRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{req.ym}_打刻一覧"

    headers = ["職種・役職", "部署", "氏名", "日付", "曜日", "種別", "時刻", "シフト予定"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font      = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align()
        c.border    = _thin()

    # ── Собираем строки ──────────────────────────────────────────────────
    rows = []
    for s in req.staff:
        dept_str = "、".join(s.departments)
        shift_str = ""
        for d in s.days:
            if d.hasShift and d.shiftStart and d.shiftEnd:
                shift_str = f"{d.shiftStart}〜{d.shiftEnd}"
            else:
                shift_str = ""

            for session in (d.sessions or []):
                entries = []
                if session.clockIn:    entries.append(("CLOCK_IN", session.clockIn))
                if session.breakStart: entries.append(("BREAK_START", session.breakStart))
                if session.breakEnd:   entries.append(("BREAK_END", session.breakEnd))
                if session.clockOut:   entries.append(("CLOCK_OUT", session.clockOut))
                if not entries:
                    continue

                for rec_type, iso_time in entries:
                    dt = _parse_dt(iso_time)
                    time_str = dt.strftime("%H:%M:%S") if dt else ""

                    # Реальная календарная дата этой конкретной пробивки
                    # (может отличаться от d.date — дня открытия смены — для ночных смен)
                    real_date_str = dt.date().isoformat() if dt else d.date
                    is_next_day   = real_date_str != d.date

                    y, m, day_num = map(int, real_date_str.split("-"))
                    wd = WD_JA[(date(y, m, day_num).weekday() + 1) % 7]

                    rows.append([
                        s.position or "", dept_str, s.userName,
                        real_date_str, wd, RECORD_LABELS.get(rec_type, rec_type),
                        time_str, shift_str, is_next_day,
                    ])

    # Сортировка: дата → 氏名 → время
    rows.sort(key=lambda row: (row[3], row[2], row[6]))

    for i, row in enumerate(rows):
        r = i + 2
        is_next_day = row[-1]
        for col, val in enumerate(row[:-1], start=1):
            c = ws.cell(row=r, column=col)
            c.value     = val
            # Дата и время пробивки, попавшей на следующий день, подсвечиваются фиолетовым
            c.font      = _font(size=9, color="7C3AED" if (col in (4, 7) and is_next_day) else "000000")
            c.alignment = _align(h="left" if col in (1, 2, 3, 8) else "center")
            c.border    = _thin()
            if i % 2 == 1:
                c.fill = _fill(C_ROW_EVEN)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()