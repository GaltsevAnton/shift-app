"""
勤怠リスト（セッション単位） — одна строка на одну смену (clock-in→clock-out).
「申請者」フィックス列 + req.visibleColumns で選択された列のみ出力
(画面の表示列トグルと同じキー)。空 = 全列表示。
"""

import io
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from models import AttendanceSessionsRequest

WD_JA       = ["日", "月", "火", "水", "木", "金", "土"]
FONT_NAME   = "メイリオ"
C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_ROW_EVEN  = "F2F2F2"
C_GRID      = "AAAAAA"

C_ACTUAL   = "009BF0"   # 実際 — rgb(0,155,240)
C_PLANNED  = "898989"   # 予定 — rgb(137,137,137)
C_NEUTRAL  = "000000"

# ── Определение опциональных колонок (соответствует LIST_COLUMNS на экране) ──
COLUMN_DEFS = {
    "scheduledInDate":  {"header": "出勤日付（予定）", "width": 16, "color": C_PLANNED, "bold": False},
    "scheduledIn":      {"header": "出勤時間（予定）", "width": 10, "color": C_PLANNED, "bold": False},
    "actualInDate":     {"header": "出勤日付（実際）", "width": 16, "color": C_ACTUAL,  "bold": False},
    "actualIn":         {"header": "出勤時間（実際）", "width": 10, "color": C_ACTUAL,  "bold": False},
    "scheduledOutDate": {"header": "退勤日付（予定）", "width": 16, "color": C_PLANNED, "bold": False},
    "scheduledOut":     {"header": "退勤時間（予定）", "width": 10, "color": C_PLANNED, "bold": False},
    "actualOutDate":    {"header": "退勤日付（実際）", "width": 16, "color": C_ACTUAL,  "bold": False},
    "actualOut":        {"header": "退勤時間（実際）", "width": 10, "color": C_ACTUAL,  "bold": False},
    "breakStart":       {"header": "休憩開始",         "width": 10, "color": C_PLANNED, "bold": False},
    "breakEnd":         {"header": "休憩終了",         "width": 10, "color": C_PLANNED, "bold": False},
    "scheduledBreak":   {"header": "休憩時間（予定）", "width": 12, "color": C_PLANNED, "bold": False},
    "actualBreakTime":  {"header": "休憩時間（実際）", "width": 12, "color": C_ACTUAL,  "bold": False},
    "workTime":         {"header": "勤務時間（予定）", "width": 12, "color": C_PLANNED, "bold": True},
    "actualWorkTime":   {"header": "勤務時間（実際）", "width": 16, "color": C_ACTUAL,  "bold": False},
    "overtimeTime":     {"header": "残業時間",         "width": 14, "color": None,      "bold": True},
    "shiftPlan":        {"header": "シフト（予定）",   "width": 14, "color": C_PLANNED, "bold": False},
}
COLUMN_ORDER = [
    "scheduledInDate", "scheduledIn", "actualInDate", "actualIn",
    "scheduledOutDate", "scheduledOut", "actualOutDate", "actualOut",
    "breakStart", "breakEnd", "scheduledBreak", "actualBreakTime",
    "workTime", "actualWorkTime", "overtimeTime", "shiftPlan",
]


def _thin(color=C_GRID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)

def _font(bold=False, color="000000", size=10, name=FONT_NAME):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _parse_dt(iso_dt):
    if not iso_dt:
        return None
    try:
        return datetime.fromisoformat(iso_dt)
    except Exception:
        return None

def _fmt_time(dt):
    return dt.strftime("%H:%M") if dt else "--:--"

def _fmt_plan_time(hhmmss):
    """LocalTime.toString() из Java — 'HH:MM' или 'HH:MM:SS'"""
    if not hhmmss:
        return "―"
    return hhmmss[:5]

def _fmt_date_wd(ds):
    try:
        y, m, d = map(int, ds.split("-"))
        wd = WD_JA[(date(y, m, d).weekday() + 1) % 7]
        return f"{ds}（{wd}）"
    except Exception:
        return ds

def _add_days_str(ds, n):
    y, m, d = map(int, ds.split("-"))
    return (date(y, m, d) + timedelta(days=n)).isoformat()

def _fmt_hm(mins):
    if mins is None:
        return "―"
    h, m = divmod(mins, 60)
    return f"{h}時間{m}分"

def _fmt_overtime(mins):
    if mins is None:
        return "―"
    sign = "+" if mins > 0 else ("-" if mins < 0 else "")
    abs_mins = abs(mins)
    h, m = divmod(abs_mins, 60)
    return f"{sign}{h}時間{m}分"

def _overtime_color(mins):
    if mins is None:
        return "94A3A8"
    if mins > 0:
        return "DC2626"
    if mins < 0:
        return "2563EB"
    return "64748B"

def _raw_break_minutes(s):
    bs, be = _parse_dt(s.breakStart), _parse_dt(s.breakEnd)
    if not bs or not be:
        return None
    mins = int((be - bs).total_seconds() / 60)
    return mins if mins > 0 else 0

def _raw_actual_worked_minutes(s):
    ci, co = _parse_dt(s.clockIn), _parse_dt(s.clockOut)
    if not ci or not co:
        return None
    mins = int((co - ci).total_seconds() / 60)
    brk = _raw_break_minutes(s)
    if brk:
        mins -= brk
    return mins if mins > 0 else 0

def _value_for(key, s):
    if key == "scheduledInDate":
        return _fmt_date_wd(s.workDate) if s.scheduledClockIn else "―"
    if key == "scheduledIn":
        return _fmt_plan_time(s.scheduledClockIn)
    if key == "actualInDate":
        ci = _parse_dt(s.clockIn)
        return _fmt_date_wd(ci.strftime("%Y-%m-%d")) if ci else "―"
    if key == "actualIn":
        return _fmt_time(_parse_dt(s.clockIn))
    if key == "scheduledOutDate":
        if not s.scheduledClockOut:
            return "―"
        ds = _add_days_str(s.workDate, 1) if s.nextDay else s.workDate
        return _fmt_date_wd(ds)
    if key == "scheduledOut":
        return _fmt_plan_time(s.scheduledClockOut)
    if key == "actualOutDate":
        co = _parse_dt(s.clockOut)
        return _fmt_date_wd(co.strftime("%Y-%m-%d")) if co else "―"
    if key == "actualOut":
        return _fmt_time(_parse_dt(s.clockOut))
    if key == "breakStart":
        bs = _parse_dt(s.breakStart)
        return _fmt_time(bs) if bs else "―"
    if key == "breakEnd":
        be = _parse_dt(s.breakEnd)
        return _fmt_time(be) if be else "―"
    if key == "scheduledBreak":
        return _fmt_hm(s.scheduledBreakMinutes)
    if key == "actualBreakTime":
        raw = _raw_break_minutes(s)
        return _fmt_hm(raw if raw is not None else s.officialBreakMinutes)
    if key == "workTime":
        return _fmt_hm(s.workMinutes)
    if key == "actualWorkTime":
        return _fmt_hm(_raw_actual_worked_minutes(s))
    if key == "overtimeTime":
        return _fmt_overtime(s.overtimeMinutes)
    if key == "shiftPlan":
        if s.scheduledClockIn and s.scheduledClockOut:
            return f"{_fmt_plan_time(s.scheduledClockIn)}〜{_fmt_plan_time(s.scheduledClockOut)}"
        return "―"
    return ""


def build(req: AttendanceSessionsRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "勤怠リスト"

    selected_keys = [k for k in COLUMN_ORDER if not req.visibleColumns or k in req.visibleColumns]

    headers = ["申請者"] + [COLUMN_DEFS[k]["header"] for k in selected_keys]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font      = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align()
        c.border    = _thin()

    # Сортировка: дата (убыв.) → 氏名
    sessions = sorted(req.sessions, key=lambda s: s.userName)
    sessions.sort(key=lambda s: s.workDate, reverse=True)

    for i, s in enumerate(sessions):
        row = [s.userName] + [_value_for(k, s) for k in selected_keys]

        r = i + 2
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col)
            c.value = val

            if col == 1:
                bold, color = False, C_NEUTRAL
            else:
                key = selected_keys[col - 2]
                bold = COLUMN_DEFS[key]["bold"]
                color = COLUMN_DEFS[key]["color"]
                if key == "overtimeTime":
                    color = _overtime_color(s.overtimeMinutes)
                elif color is None:
                    color = C_NEUTRAL

            c.font      = _font(bold=bold, color=color)
            c.alignment = _align(h="left" if col == 1 else "center")
            c.border    = _thin()
            if i % 2 == 1:
                c.fill = _fill(C_ROW_EVEN)

    widths = [14] + [COLUMN_DEFS[k]["width"] for k in selected_keys]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(sessions) + 1}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()