"""
Табель учёта фактического времени (勤怠集計表・実績).
Каждая смена сотрудника за день — отдельный блок из 4 строк (出勤/退勤/実働/休憩),
блоки разделены жирной линией — как в календарном виде 勤怠管理 на экране.
Высота блока сотрудника = 4 × (макс. число смен в любом дне месяца у этого сотрудника).
実働・休憩 считаются гибридно (факт пробивки休憩, иначе — авто по 休憩ルール) —
все вычисления уже сделаны на бэкенде (ReportService), билдер только форматирует.
Колонки: 職種・役職 | 部署 | 氏名 | メタ | 1..31 | 合計時間
"""

import io
import calendar
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from models import AttendanceReportRequest, AttendanceStaffModel, AttendanceDayModel

WD_JA     = ["日", "月", "火", "水", "木", "金", "土"]
FONT_NAME = "メイリオ"

C_SAT_BG     = "BDD7EE"
C_SUN_BG     = "FCE4D6"
C_WEEKDAY_BG = "DDEBF7"
C_OFF_FG     = "808080"
C_OFF_BG     = "F2F2F2"
C_ROW_ODD    = "FFFFFF"
C_LABEL_BG   = "D9D9D9"
C_GRID       = "AAAAAA"
C_LATE_BG    = "FEE2E2"   # опоздание — красноватый

LABELS = ["出勤", "退勤", "実働", "休憩"]


def _thin(color=C_GRID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)

def _font(bold=False, color="000000", size=9, name=FONT_NAME):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _weekday(ym, day):
    y, m = map(int, ym.split("-"))
    wd = date(y, m, day).weekday()
    return (wd + 1) % 7  # 0=日..6=土

def _day_data(staff: AttendanceStaffModel, day: int, ym: str) -> AttendanceDayModel:
    ds = f"{ym}-{day:02d}"
    for d in staff.days:
        if d.date == ds:
            return d
    return AttendanceDayModel(date=ds)

def _parse_hm(iso_dt: str | None):
    if not iso_dt:
        return None
    try:
        return datetime.fromisoformat(iso_dt).strftime("%H:%M")
    except Exception:
        return None

def _fmt_hm(mins):
    if mins is None:
        return ""
    h, m = divmod(mins, 60)
    return f"{h}:{m:02d}"

def _apply_outer_border(ws, min_row, max_row, min_col, max_col, color=C_GRID):
    thin = Side(style="thin", color=color)
    no   = Side(style=None)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thin if row == min_row else no
            bottom = thin if row == max_row else no
            left   = thin if col == min_col else no
            right  = thin if col == max_col else no
            c.border = Border(top=top, bottom=bottom, left=left, right=right)

def _max_sessions_for_staff(s: AttendanceStaffModel, days, ym: str) -> int:
    m = 1
    for day in days:
        cnt = len(_day_data(s, day, ym).sessions or [])
        if cnt > m:
            m = cnt
    return m


def build(req: AttendanceReportRequest) -> bytes:
    ym    = req.ym
    y, m  = map(int, ym.split("-"))
    total = calendar.monthrange(y, m)[1]
    days  = list(range(1, total + 1))
    staff = req.staff

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}年{m}月_勤怠集計"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 5
    for i, _ in enumerate(days):
        col = get_column_letter(5 + i)
        ws.column_dimensions[col].width = 7
    ws.column_dimensions[get_column_letter(5 + total)].width = 7  # 合計時間

    ws.row_dimensions[1].height = 6

    last_col = get_column_letter(4 + total)
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"{req.hotelName}勤怠集計表（実績）（{y}年{m}月1日～{m}月{total}日）"
    c.font      = _font(bold=True, size=12)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 5

    for col_idx, label in [(1, "職種・役職"), (2, "部署"), (3, "氏名"), (4, "日\n曜日")]:
        ws.merge_cells(f"{get_column_letter(col_idx)}4:{get_column_letter(col_idx)}5")
        c = ws.cell(row=4, column=col_idx)
        c.value     = label
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()
        _apply_outer_border(ws, 4, 5, col_idx, col_idx)

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16

    for i, day in enumerate(days):
        col = 5 + i
        wd  = _weekday(ym, day)
        c = ws.cell(row=4, column=col); c.value = day; c.font = _font(size=8); c.alignment = _align(); c.border = _thin()
        c = ws.cell(row=5, column=col); c.value = WD_JA[wd]; c.font = _font(size=8); c.alignment = _align(); c.border = _thin()

    ws.merge_cells(f"{get_column_letter(5 + total)}4:{get_column_letter(5 + total)}5")
    c = ws.cell(row=4, column=5 + total)
    c.value     = "合計\n時間"
    c.font      = _font(size=7)
    c.alignment = _align()
    c.border    = _thin()
    _apply_outer_border(ws, 4, 5, 5 + total, 5 + total)

    base_row = 6
    current_row = base_row

    for staff_idx, s in enumerate(staff):
        max_sessions = _max_sessions_for_staff(s, days, ym)
        rows_per_staff = 4 * max_sessions
        r = current_row
        row_bg = C_ROW_ODD

        ws.merge_cells(f"A{r}:A{r+rows_per_staff-1}")
        c = ws[f"A{r}"]; c.value = s.position or ""; c.font = _font(size=8, color="555555")
        c.fill = _fill(row_bg); c.alignment = _align(h="center"); c.border = _thin()
        _apply_outer_border(ws, r, r+rows_per_staff-1, 1, 1)

        ws.merge_cells(f"B{r}:B{r+rows_per_staff-1}")
        c = ws[f"B{r}"]; c.value = "、".join(s.departments); c.font = _font(size=8, color="555555")
        c.fill = _fill(row_bg); c.alignment = _align(h="center"); c.border = _thin()
        _apply_outer_border(ws, r, r+rows_per_staff-1, 2, 2)

        ws.merge_cells(f"C{r}:C{r+rows_per_staff-1}")
        c = ws[f"C{r}"]; c.value = s.userName; c.font = _font(bold=True, size=9)
        c.fill = _fill(row_bg); c.alignment = _align(h="center"); c.border = _thin()
        _apply_outer_border(ws, r, r+rows_per_staff-1, 3, 3)

        # Метки 出勤/退勤/実働/休憩 — повторяются на каждый блок-смену
        for session_idx in range(max_sessions):
            block_r = r + session_idx * 4
            for sub, label in enumerate(LABELS):
                c = ws.cell(row=block_r + sub, column=4)
                c.value = label; c.font = _font(size=7, color="333333")
                c.fill = _fill(C_LABEL_BG); c.alignment = _align(); c.border = _thin()
                ws.row_dimensions[block_r + sub].height = 14

        total_work_minutes = 0
        for i, day in enumerate(days):
            col = 5 + i
            d = _day_data(s, day, ym)
            sessions = d.sessions or []

            if not sessions:
                ws.merge_cells(f"{get_column_letter(col)}{r}:{get_column_letter(col)}{r+rows_per_staff-1}")
                c = ws.cell(row=r, column=col)
                c.value = "―"; c.font = _font(color=C_OFF_FG, size=10)
                c.fill = _fill(C_OFF_BG); c.alignment = _align(); c.border = _thin()
                _apply_outer_border(ws, r, r+rows_per_staff-1, col, col)
                continue

            day_work_min = sum(sess.workMinutes or 0 for sess in sessions)
            total_work_minutes += day_work_min

            for session_idx in range(max_sessions):
                block_r = r + session_idx * 4
                sess = sessions[session_idx] if session_idx < len(sessions) else None

                if sess is None:
                    # У этого дня меньше смен, чем максимум по сотруднику — оставляем пусто
                    for sub in range(4):
                        c = ws.cell(row=block_r + sub, column=col)
                        c.value = ""
                        c.fill = _fill(row_bg)
                        c.alignment = _align(h="center", v="center")
                        c.border = _thin()
                        c.font = _font(size=9)
                    continue

                in_time    = _parse_hm(sess.officialClockIn) or ""
                out_time   = _parse_hm(sess.officialClockOut) or ""
                work_time  = _fmt_hm(sess.workMinutes)
                break_time = _fmt_hm(sess.officialBreakMinutes)
                block_bg   = C_LATE_BG if (sess.lateIn or sess.earlyOut) else row_bg

                for sub, val in enumerate([in_time, out_time, work_time, break_time]):
                    c = ws.cell(row=block_r + sub, column=col)
                    c.value = val; c.fill = _fill(block_bg)
                    c.alignment = _align(h="center", v="center"); c.border = _thin()
                    if sub == 2:
                        c.font = _font(size=9, bold=True, color="0369A1")
                    elif sub == 3:
                        c.font = _font(size=8, color="777777")
                    else:
                        c.font = _font(size=9)

        off_col = 5 + total
        ws.merge_cells(f"{get_column_letter(off_col)}{r}:{get_column_letter(off_col)}{r+rows_per_staff-1}")
        c = ws.cell(row=r, column=off_col)
        c.value = _fmt_hm(total_work_minutes); c.font = _font(bold=True, size=9)
        c.alignment = _align(); c.border = _thin()
        _apply_outer_border(ws, r, r+rows_per_staff-1, off_col, off_col)

        # Разделитель между блоками смен (жирная линия) + внешняя рамка строки сотрудника
        thick = Side(style="medium", color="000000")
        no    = Side(style=None)
        is_first = staff_idx == 0
        for sub_row in range(r, r + rows_per_staff):
            is_session_divider = (sub_row - r + 1) % 4 == 0 and sub_row != r + rows_per_staff - 1
            for col_idx in range(1, 5 + total):
                c = ws.cell(row=sub_row, column=col_idx)
                top    = thick if (sub_row == r and is_first) else no
                bottom = thick if (sub_row == r + rows_per_staff - 1 or is_session_divider) else no
                left   = thick if col_idx == 1 else no
                right  = thick if col_idx == 4 + total else no
                if any([top != no, bottom != no, left != no, right != no]):
                    existing = c.border
                    c.border = Border(
                        top=top if top != no else existing.top,
                        bottom=bottom if bottom != no else existing.bottom,
                        left=left if left != no else existing.left,
                        right=right if right != no else existing.right,
                    )

        current_row += rows_per_staff

    last_data_row = current_row - 1
    last_data_col = 5 + total
    thick = Side(style="medium", color="000000")
    no    = Side(style=None)
    for row in range(4, last_data_row + 1):
        for col in range(1, last_data_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thick if row == 4 else no
            bottom = thick if row == last_data_row else no
            left   = thick if col == 1 else no
            right  = thick if col == last_data_col else no
            if any([top, bottom, left, right]):
                existing = c.border
                c.border = Border(
                    top=top if top != no else existing.top,
                    bottom=bottom if bottom != no else existing.bottom,
                    left=left if left != no else existing.left,
                    right=right if right != no else existing.right,
                )

    ws.freeze_panes = "E6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    ws.page_setup.copies = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()