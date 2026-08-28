"""
Сводная шифт-таблица по всем сотрудникам.
Формат как shift_dept: 3 строки на сотрудника (出勤/退勤/職場).
Колонки: 職種・役職 | 部署 | 氏名 | メタ | 1..31 | 公休数
"""

import io
import calendar
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from models import ReportRequest, ReportRangeRequest, StaffModel, DayModel

WD_JA     = ["日", "月", "火", "水", "木", "金", "土"]
FONT_NAME = "メイリオ"

C_HEADER_BG  = "1F4E79"
C_HEADER_FG  = "FFFFFF"
C_SAT_BG     = "BDD7EE"
C_SUN_BG     = "FCE4D6"
C_WEEKDAY_BG = "DDEBF7"
C_OFF_FG     = "808080"
C_OFF_BG     = "F2F2F2"
C_ROW_ODD    = "FFFFFF"
C_ROW_EVEN   = "F2F2F2"
C_LABEL_BG   = "D9D9D9"
C_GRID       = "AAAAAA"


def _thin(color=C_GRID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)

def _font(bold=False, color="000000", size=9, name=FONT_NAME):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _weekday(ym, day):
    y, m = map(int, ym.split("-"))
    wd = date(y, m, day).weekday()
    return (wd + 1) % 7  # 0=日..6=土

def _day_data(staff: StaffModel, day: int, ym: str) -> DayModel:
    ds = f"{ym}-{day:02d}"
    for d in staff.days:
        if d.date == ds:
            return d
    return DayModel(date=ds, off=True, slots=[])

def _fmt(t): return t[:5] if t else ""

def _apply_outer_border(ws, min_row, max_row, min_col, max_col, color=C_GRID):
    thin = Side(style="thin", color=color)
    no   = Side(style=None)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thin if row == min_row else no
            bottom = thin if row == max_row else no
            left   = thin if col == min_col else no
            right  = thin if col == max_col else no
            c.border = Border(top=top, bottom=bottom, left=left, right=right)

def _date_range(from_str: str, to_str: str) -> list[str]:
    y1, m1, d1 = map(int, from_str.split("-"))
    y2, m2, d2 = map(int, to_str.split("-"))
    cur = date(y1, m1, d1)
    end = date(y2, m2, d2)
    out = []
    while cur <= end:
        out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out

def _weekday_from_iso(date_str: str) -> int:
    y, m, d = map(int, date_str.split("-"))
    wd = date(y, m, d).weekday()
    return (wd + 1) % 7

def _day_data_range(staff: StaffModel, date_str: str) -> DayModel:
    for d in staff.days:
        if d.date == date_str:
            return d
    return DayModel(date=date_str, off=True, slots=[])

def build(req: ReportRequest) -> bytes:
    ym    = req.ym
    y, m  = map(int, ym.split("-"))
    total = calendar.monthrange(y, m)[1]
    days  = list(range(1, total + 1))
    staff = req.staff

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}年{m}月_全員"

    # Колонки: A=職種, B=部署, C=氏名, D=メタ(出勤/退勤/職場), E..=дни, последняя=公休数
    # ── Ширина колонок ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6     # 職種
    ws.column_dimensions["B"].width = 8     # 部署
    ws.column_dimensions["C"].width = 12    # 氏名
    ws.column_dimensions["D"].width = 5     # メタ
    for i, _ in enumerate(days):
        col = get_column_letter(5 + i)
        ws.column_dimensions[col].width = 6
    ws.column_dimensions[get_column_letter(5 + total)].width = 5  # 公休数

    # ── Строка 1: пусто ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6

    # ── Строка 2: заголовок ──────────────────────────────────────────────
    last_col = get_column_letter(4 + total)
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"{req.hotelName}シフト表（{y}年{m}月1日～{m}月{total}日） 全員"
    c.font      = _font(bold=True, size=12)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 18

    # ── Строка 3: пусто ──────────────────────────────────────────────────
    ws.row_dimensions[3].height = 5

    # ── Строки 4-5: заголовки дней ───────────────────────────────────────
    # Merge A4:A5, B4:B5, C4:C5, D4:D5
    for col_idx, label in [(1, "職種・役職"), (2, "部署"), (3, "氏名"), (4, "日\n曜日")]:
        ws.merge_cells(f"{get_column_letter(col_idx)}4:{get_column_letter(col_idx)}5")
        c = ws.cell(row=4, column=col_idx)
        c.value     = label
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()
        _apply_outer_border(ws, 4, 5, col_idx, col_idx)

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16

    # Числа и дни недели
    for i, day in enumerate(days):
        col = 5 + i
        wd  = _weekday(ym, day)

        c = ws.cell(row=4, column=col)
        c.value     = day
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()

        c = ws.cell(row=5, column=col)
        c.value     = WD_JA[wd]
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()

    # Заголовок 公休数
    ws.merge_cells(f"{get_column_letter(5 + total)}4:{get_column_letter(5 + total)}5")
    c = ws.cell(row=4, column=5 + total)
    c.value     = "公休\n数"
    c.font      = _font(size=7)
    c.alignment = _align()
    c.border    = _thin()
    _apply_outer_border(ws, 4, 5, 5 + total, 5 + total)

    # ── Данные сотрудников ────────────────────────────────────────────────
    base_row = 6
    for staff_idx, s in enumerate(staff):
        r      = base_row + staff_idx * 3
        row_bg = C_ROW_ODD

        # Merge A(r..r+2) — 職種
        ws.merge_cells(f"A{r}:A{r+2}")
        c = ws[f"A{r}"]
        c.value     = s.position or ""
        c.font      = _font(size=8, color="555555")
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 1, 1)

        # Merge B(r..r+2) — 部署
        ws.merge_cells(f"B{r}:B{r+2}")
        c = ws[f"B{r}"]
        c.value     = "、".join(s.departments)
        c.font      = _font(size=8, color="555555")
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 2, 2)

        # Merge C(r..r+2) — 氏名
        ws.merge_cells(f"C{r}:C{r+2}")
        c = ws[f"C{r}"]
        c.value     = s.userName
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 3, 3)

        # D labels: 出勤 / 退勤 / 職場
        for sub, label in enumerate(["出勤", "退勤", "職場"]):
            c = ws.cell(row=r + sub, column=4)
            c.value     = label
            c.font      = _font(size=7, color="333333")
            c.fill      = _fill(C_LABEL_BG)
            c.alignment = _align()
            c.border    = _thin()
            ws.row_dimensions[r + sub].height = 14

        # Данные по дням
        for i, day in enumerate(days):
            col = 5 + i
            day_bg = row_bg

            d = _day_data(s, day, ym)

            if d.off or not d.slots:
                ws.merge_cells(f"{get_column_letter(col)}{r}:{get_column_letter(col)}{r+2}")
                c = ws.cell(row=r, column=col)
                c.value     = "休"
                c.font      = _font(color=C_OFF_FG, bold=True, size=10)
                c.fill      = _fill(C_OFF_BG)
                c.alignment = _align()
                c.border    = _thin()
                _apply_outer_border(ws, r, r+2, col, col)
            else:
                slots = d.slots
                starts    = "\n".join(_fmt(sl.startTime) for sl in slots if sl.startTime)
                ends      = "\n".join("L" if sl.last else _fmt(sl.endTime) for sl in slots)
                workplaces = "\n".join(sl.workplace or "" for sl in slots)

                for sub, val in enumerate([starts, ends, workplaces]):
                    c = ws.cell(row=r + sub, column=col)
                    c.value     = val
                    c.fill      = _fill(day_bg)
                    c.alignment = _align(h="center", v="center")
                    c.border    = _thin()
                    c.font      = _font(size=9 if sub in (0, 1) else 8,
                                        color="555555" if sub == 2 else "000000")

        # 公休数
        off_count = sum(
            1 for day_num in days
            if not _day_data(s, day_num, ym).slots or _day_data(s, day_num, ym).off
        )
        off_col = 5 + total
        ws.merge_cells(f"{get_column_letter(off_col)}{r}:{get_column_letter(off_col)}{r+2}")
        c = ws.cell(row=r, column=off_col)
        c.value     = off_count
        c.font      = _font(size=9)
        c.alignment = _align()
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, off_col, off_col)

        # Внешняя рамка строки
        thick = Side(style="medium", color="000000")
        no    = Side(style=None)
        is_first = staff_idx == 0

        for sub_row in range(r, r + 3):
            for col_idx in range(1, 5 + total):
                c = ws.cell(row=sub_row, column=col_idx)
                top    = thick if (sub_row == r and is_first) else no
                bottom = thick if sub_row == r + 2 else no
                left   = thick if col_idx == 1 else no
                right  = thick if col_idx == 4 + total else no
                if any([top != no, bottom != no, left != no, right != no]):
                    existing = c.border
                    c.border = Border(
                        top    = top    if top    != no else existing.top,
                        bottom = bottom if bottom != no else existing.bottom,
                        left   = left   if left   != no else existing.left,
                        right  = right  if right  != no else existing.right,
                    )

    # Внешняя рамка всей таблицы
    last_data_row = base_row + len(staff) * 3 - 1
    last_data_col = 5 + total

    thick = Side(style="medium", color="000000")
    no    = Side(style=None)

    for row in range(4, last_data_row + 1):
        for col in range(1, last_data_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thick if row == 4             else no
            bottom = thick if row == last_data_row else no
            left   = thick if col == 1             else no
            right  = thick if col == last_data_col else no
            if any([top, bottom, left, right]):
                existing = c.border
                c.border = Border(
                    top    = top    if top    != no else existing.top,
                    bottom = bottom if bottom != no else existing.bottom,
                    left   = left   if left   != no else existing.left,
                    right  = right  if right  != no else existing.right,
                )

    ws.freeze_panes = "E6"

    # Настройки печати
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    ws.page_setup.copies = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def build_range(req: ReportRangeRequest) -> bytes:
    dates = _date_range(req.fromDate, req.toDate)
    total = len(dates)
    staff = req.staff

    wb = Workbook()
    ws = wb.active
    ws.title = f"{req.fromDate}〜{req.toDate}_全員"[:31]

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 5
    for i, _ in enumerate(dates):
        col = get_column_letter(5 + i)
        ws.column_dimensions[col].width = 6
    ws.column_dimensions[get_column_letter(5 + total)].width = 5

    ws.row_dimensions[1].height = 6

    last_col = get_column_letter(4 + total)
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"{req.hotelName}シフト表（{req.fromDate}～{req.toDate}） 全員"
    c.font      = _font(bold=True, size=12)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 5

    for col_idx, label in [(1, "職種・役職"), (2, "部署"), (3, "氏名"), (4, "日\n曜日")]:
        ws.merge_cells(f"{get_column_letter(col_idx)}4:{get_column_letter(col_idx)}5")
        c = ws.cell(row=4, column=col_idx)
        c.value     = label
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()
        _apply_outer_border(ws, 4, 5, col_idx, col_idx)

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16

    for i, ds in enumerate(dates):
        col = 5 + i
        wd  = _weekday_from_iso(ds)
        mm, dd = int(ds[5:7]), int(ds[8:10])

        c = ws.cell(row=4, column=col)
        c.value     = f"{mm}/{dd}"
        c.font      = _font(size=7)
        c.alignment = _align()
        c.border    = _thin()

        c = ws.cell(row=5, column=col)
        c.value     = WD_JA[wd]
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()

    ws.merge_cells(f"{get_column_letter(5 + total)}4:{get_column_letter(5 + total)}5")
    c = ws.cell(row=4, column=5 + total)
    c.value     = "公休\n数"
    c.font      = _font(size=7)
    c.alignment = _align()
    c.border    = _thin()
    _apply_outer_border(ws, 4, 5, 5 + total, 5 + total)

    base_row = 6
    for staff_idx, s in enumerate(staff):
        r      = base_row + staff_idx * 3
        row_bg = C_ROW_ODD

        ws.merge_cells(f"A{r}:A{r+2}")
        c = ws[f"A{r}"]
        c.value     = s.position or ""
        c.font      = _font(size=8, color="555555")
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 1, 1)

        ws.merge_cells(f"B{r}:B{r+2}")
        c = ws[f"B{r}"]
        c.value     = "、".join(s.departments)
        c.font      = _font(size=8, color="555555")
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 2, 2)

        ws.merge_cells(f"C{r}:C{r+2}")
        c = ws[f"C{r}"]
        c.value     = s.userName
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 3, 3)

        for sub, label in enumerate(["出勤", "退勤", "職場"]):
            c = ws.cell(row=r + sub, column=4)
            c.value     = label
            c.font      = _font(size=7, color="333333")
            c.fill      = _fill(C_LABEL_BG)
            c.alignment = _align()
            c.border    = _thin()
            ws.row_dimensions[r + sub].height = 14

        for i, ds in enumerate(dates):
            col = 5 + i
            day_bg = row_bg

            d = _day_data_range(s, ds)

            if d.off or not d.slots:
                ws.merge_cells(f"{get_column_letter(col)}{r}:{get_column_letter(col)}{r+2}")
                c = ws.cell(row=r, column=col)
                c.value     = "休"
                c.font      = _font(color=C_OFF_FG, bold=True, size=10)
                c.fill      = _fill(C_OFF_BG)
                c.alignment = _align()
                c.border    = _thin()
                _apply_outer_border(ws, r, r+2, col, col)
            else:
                slots = d.slots
                starts     = "\n".join(_fmt(sl.startTime) for sl in slots if sl.startTime)
                ends       = "\n".join("L" if sl.last else _fmt(sl.endTime) for sl in slots)
                workplaces = "\n".join(sl.workplace or "" for sl in slots)

                for sub, val in enumerate([starts, ends, workplaces]):
                    c = ws.cell(row=r + sub, column=col)
                    c.value     = val
                    c.fill      = _fill(day_bg)
                    c.alignment = _align(h="center", v="center")
                    c.border    = _thin()
                    c.font      = _font(size=9 if sub in (0, 1) else 8,
                                        color="555555" if sub == 2 else "000000")

        off_count = sum(1 for ds in dates if not _day_data_range(s, ds).slots or _day_data_range(s, ds).off)
        off_col = 5 + total
        ws.merge_cells(f"{get_column_letter(off_col)}{r}:{get_column_letter(off_col)}{r+2}")
        c = ws.cell(row=r, column=off_col)
        c.value     = off_count
        c.font      = _font(size=9)
        c.alignment = _align()
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, off_col, off_col)

        thick = Side(style="medium", color="000000")
        no    = Side(style=None)
        is_first = staff_idx == 0

        for sub_row in range(r, r + 3):
            for col_idx in range(1, 5 + total):
                c = ws.cell(row=sub_row, column=col_idx)
                top    = thick if (sub_row == r and is_first) else no
                bottom = thick if sub_row == r + 2 else no
                left   = thick if col_idx == 1 else no
                right  = thick if col_idx == 4 + total else no
                if any([top != no, bottom != no, left != no, right != no]):
                    existing = c.border
                    c.border = Border(
                        top    = top    if top    != no else existing.top,
                        bottom = bottom if bottom != no else existing.bottom,
                        left   = left   if left   != no else existing.left,
                        right  = right  if right  != no else existing.right,
                    )

    last_data_row = base_row + len(staff) * 3 - 1
    last_data_col = 5 + total

    thick = Side(style="medium", color="000000")
    no    = Side(style=None)

    for row in range(4, last_data_row + 1):
        for col in range(1, last_data_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thick if row == 4             else no
            bottom = thick if row == last_data_row else no
            left   = thick if col == 1             else no
            right  = thick if col == last_data_col else no
            if any([top, bottom, left, right]):
                existing = c.border
                c.border = Border(
                    top    = top    if top    != no else existing.top,
                    bottom = bottom if bottom != no else existing.bottom,
                    left   = left   if left   != no else existing.left,
                    right  = right  if right  != no else existing.right,
                )

    ws.freeze_panes = "E6"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    ws.page_setup.copies = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()