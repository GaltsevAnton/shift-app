"""
Шифт-таблица по отделу.
Формат точно как образец シフト_2026-05-洋食.xlsx:

  Строка 1: [пусто]
  Строка 2: ホテル・ヘリテイジ飯能sta．シフト表（YYYY年M月1日～M月DD日）
  Строка 3: 【M月シフト {department}】
  Строка 4: [пусто]
  Строка 5: 役職 | 氏名 | 日  | 1 | 2 | ... | 31
  Строка 6:       |      | 曜日| 金| 土 | ... | 日
  На каждого сотрудника 4 строки:
    R+0: 役職(merge R..R+2) | 氏名(merge) | 出勤 | start1\nstart2...
    R+1:                    |             | 退勤 | end1\nend2...
    R+2:                    |             | 職場 | 
"""

import io
import calendar
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from models import ReportRequest, StaffModel, DayModel, SlotModel, BreakRuleModel
from openpyxl.worksheet.page import PageMargins

# ── Константы ──────────────────────────────────────────────────────────────
WD_JA   = ["日", "月", "火", "水", "木", "金", "土"]
FONT_NAME = "メイリオ"

# Цвета
C_HEADER_BG   = "1F4E79"   # тёмно-синий — шапка 役職/氏名/日
C_HEADER_FG   = "FFFFFF"
C_SAT_BG      = "BDD7EE"   # голубой — суббота
C_SUN_BG      = "FCE4D6"   # розовый — воскресенье
C_WEEKDAY_BG  = "DDEBF7"   # светло-голубой — будни заголовок
C_OFF_FG      = "808080"   # красный — 休
C_OFF_BG      = "F2F2F2"
C_ROW_ODD     = "FFFFFF"
C_ROW_EVEN    = "F2F2F2"
C_LABEL_BG    = "D9D9D9"   # серый — ячейки 出勤/退勤/職場
C_GRID        = "AAAAAA"


def _thin(color=C_GRID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=9, name=FONT_NAME):
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _days_in_month(ym: str):
    y, m = map(int, ym.split("-"))
    return calendar.monthrange(y, m)[1]


def _weekday(ym: str, day: int) -> int:
    """0=月...6=日  →  конвертируем в 0=日..6=土"""
    y, m = map(int, ym.split("-"))
    wd = date(y, m, day).weekday()  # 0=Mon..6=Sun
    return (wd + 1) % 7             # 0=日,1=月,...,6=土


def _day_data(staff: StaffModel, day: int, ym: str) -> DayModel:
    ds = f"{ym}-{day:02d}"
    for d in staff.days:
        if d.date == ds:
            return d
    return DayModel(date=ds, off=True, slots=[])


def _format_time(t: str | None) -> str:
    if not t:
        return ""
    return t[:5]
def _slot_duration_minutes(sl: SlotModel):
    if not sl.startTime or not sl.endTime:
        return None
    try:
        sh, sm = map(int, sl.startTime[:5].split(":"))
        eh, em = map(int, sl.endTime[:5].split(":"))
        start_min = sh * 60 + sm
        end_min   = eh * 60 + em
        if end_min <= start_min:
            end_min += 24 * 60
        return end_min - start_min
    except Exception:
        return None

def _auto_break_minutes(duration_min, break_rules: list[BreakRuleModel]):
    if duration_min is None:
        return None
    applicable = [r for r in break_rules if duration_min > r.thresholdMinutes]
    if not applicable:
        return 0
    best = max(applicable, key=lambda r: r.thresholdMinutes)
    return best.breakMinutes

def _fmt_break(mins):
    if mins is None:
        return ""
    h, m = divmod(mins, 60)
    return f"{h}:{m:02d}"

def _apply_outer_border(ws, min_row, max_row, min_col, max_col, color=C_GRID):
    """Рисует внешнюю рамку вокруг диапазона ячеек."""
    thin = Side(style="thin", color=color)
    no   = Side(style=None)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thin if row == min_row else no
            bottom = thin if row == max_row else no
            left   = thin if col == min_col else no
            right  = thin if col == max_col else no
            c.border = Border(top=top, bottom=bottom, left=left, right=right)

def _medium(color=C_GRID):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def build(req: ReportRequest) -> bytes:
    ym       = req.ym
    y, m     = map(int, ym.split("-"))
    total    = _days_in_month(ym)
    days     = list(range(1, total + 1))
    staff    = req.staff

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}年{m}月"


    # ── Ширина колонок ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5.25    # 役職
    ws.column_dimensions["B"].width = 12   # 氏名
    ws.column_dimensions["C"].width = 5    # 日/曜日/出勤/退勤/職場
    for i, _ in enumerate(days):
        col = get_column_letter(4 + i)
        ws.column_dimensions[col].width = 6
    ws.column_dimensions[get_column_letter(4 + total)].width = 5

    # ── Строка 1: пусто ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6

    # ── Строка 2: заголовок отеля ─────────────────────────────────────────
    last_col = get_column_letter(3 + total)
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (
        f"{req.hotelName}シフト表"
        f"（{y}年{m}月1日～{m}月{total}日）"
    )
    c.font      = _font(bold=True, size=12)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 18

    # ── Строка 3: название отдела ─────────────────────────────────────────
    ws.merge_cells(f"B3:{last_col}3")
    c = ws["B3"]
    c.value     = f"【{m}月シフト {req.department or ''}】"
    c.font      = _font(bold=True, size=12)
    c.alignment = _align(h="left")
    ws.row_dimensions[3].height = 16

    # ── Строка 4: пусто ───────────────────────────────────────────────────
    ws.row_dimensions[4].height = 5

    # ── Строки 5-6: заголовки дней ───────────────────────────────────────
    # Merge A5:A6 (役職), B5:B6 (氏名)
    ws.merge_cells("A5:A6")
    ws.merge_cells("B5:B6")
    _apply_outer_border(ws, 5, 6, 1, 1)  # A5:A6
    _apply_outer_border(ws, 5, 6, 2, 2)  # B5:B6

    for cell_addr, label in [("A5", "役職"), ("B5", "氏名")]:
        c = ws[cell_addr]
        c.value     = label
        c.font      = _font(size=9)
        c.alignment = _align()
        c.border    = _thin()

    # C5 = 日, C6 = 曜日
    for r, label in [(5, "日"), (6, "曜日")]:
        c = ws.cell(row=r, column=3)
        c.value     = label
        c.font      = _font(size=7) 
        c.alignment = _align()
        c.border    = _thin()

    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 16

    # Числа и дни недели
    for i, day in enumerate(days):
        col = 4 + i
        wd  = _weekday(ym, day)
        is_sat = wd == 6
        is_sun = wd == 0
        bg = C_SAT_BG if is_sat else (C_SUN_BG if is_sun else C_WEEKDAY_BG)

        # Строка 5: число
        c = ws.cell(row=5, column=col)
        c.value     = day
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()

        # Строка 6: день недели
        c = ws.cell(row=6, column=col)
        c.value     = WD_JA[wd]
        c.font      = _font(size=8)
        c.alignment = _align()
        c.border    = _thin()

    # ── Заголовок 公休数 ──
    ws.merge_cells(f"{get_column_letter(4 + total)}5:{get_column_letter(4 + total)}6")
    c = ws.cell(row=5, column=4 + total)
    c.value     = "公休\n数"
    c.font      = _font(size=7)
    c.alignment = _align()
    c.border    = _thin()
    _apply_outer_border(ws, 5, 6, 4 + total, 4 + total)

    # ── Данные сотрудников ────────────────────────────────────────────────
    base_row = 7
    for staff_idx, s in enumerate(staff):
        r = base_row + staff_idx * 3
        row_bg = C_ROW_ODD

        # Merge A(r..2) — 役職
        ws.merge_cells(f"A{r}:A{r+2}")
        c = ws[f"A{r}"]
        c.value     = s.position or ""
        c.font      = _font(size=8, color="555555")
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 1, 1)

        # Merge B(r..r+2) — 氏名
        ws.merge_cells(f"B{r}:B{r+2}")
        c = ws[f"B{r}"]
        c.value     = s.userName
        c.font      = _font(size=9)
        c.fill      = _fill(row_bg)
        c.alignment = _align(h="center")
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, 2, 2)

        # C labels: 出勤 / 退勤 / 休憩
        labels = ["出勤", "退勤", "休憩"]
        for sub, label in enumerate(labels):
            c = ws.cell(row=r + sub, column=3)
            c.value     = label
            c.font      = _font(size=7, color="333333")
            c.fill      = _fill(C_LABEL_BG)
            c.alignment = _align()
            c.border    = _thin()
            ws.row_dimensions[r + sub].height = 14

        # Данные по дням
        for i, day in enumerate(days):
            col = 4 + i
            wd  = _weekday(ym, day)
            is_sat = wd == 6
            is_sun = wd == 0
            day_bg = row_bg

            d = _day_data(s, day, ym)

            if d.off or not d.slots:
                # Merge всех 4 строк для 休
                ws.merge_cells(f"{get_column_letter(col)}{r}:{get_column_letter(col)}{r+2}")
                c = ws.cell(row=r, column=col)
                c.value     = "休"
                c.font      = _font(color=C_OFF_FG, bold=True, size=10)
                c.fill      = _fill(C_OFF_BG)
                c.alignment = _align()
                c.border    = _thin()
                _apply_outer_border(ws, r, r+2, col, col)
            else:
                slots = d.slots
                # Строка 出勤 (r+0): начало через \n
                starts = "\n".join(
                    _format_time(sl.startTime) for sl in slots if sl.startTime
                )
                # Строка 退勤 (r+1): конец через \n, L если last
                ends = "\n".join(
                    ("L" if sl.last else _format_time(sl.endTime)) for sl in slots
                )
                # Строка 休憩 (r+2): авто по 休憩ルール от длительности слота
                breaks = "\n".join(
                    _fmt_break(_auto_break_minutes(_slot_duration_minutes(sl), req.breakRules))
                    for sl in slots
                )

                for sub, (val, is_label_row) in enumerate([
                    (starts, False),
                    (ends,   False),
                    (breaks, False),
                ]):
                    c = ws.cell(row=r + sub, column=col)
                    c.value     = val
                    c.fill      = _fill(day_bg)
                    c.alignment = _align(h="center", v="center")
                    c.border    = _thin()
                    if sub == 3:
                        c.font = _font(size=8, color="555555")
                    elif sub in (0, 1):
                        c.font = _font(size=9)
                    else:
                        c.font = _font(size=8)

        # ── Ячейка 公休数 ──
        off_count = sum(
            1 for day_num in days
            if not _day_data(s, day_num, ym).slots or _day_data(s, day_num, ym).off
        )
        off_col = 4 + total
        ws.merge_cells(f"{get_column_letter(off_col)}{r}:{get_column_letter(off_col)}{r+2}")
        c = ws.cell(row=r, column=off_col)
        c.value     = off_count
        c.font      = _font(size=9)
        c.alignment = _align()
        c.border    = _thin()
        _apply_outer_border(ws, r, r+2, off_col, off_col)

        # ── Внешняя рамка строки сотрудника ──────────────────────────────────
        thick = Side(style="medium", color="000000")
        no    = Side(style=None)
        is_first = staff_idx == 0
        is_last  = staff_idx == len(staff) - 1

        for sub_row in range(r, r + 3):
            for col_idx in range(1, 4 + total):
                c = ws.cell(row=sub_row, column=col_idx)
                top    = thick if (sub_row == r and is_first) else no
                bottom = thick if sub_row == r + 2 else no
                left   = thick if col_idx == 1 else no
                right  = thick if col_idx == 3 + total else no
                if any([top != no, bottom != no, left != no, right != no]):
                    existing = c.border
                    c.border = Border(
                        top    = top    if top    != no else existing.top,
                        bottom = bottom if bottom != no else existing.bottom,
                        left   = left   if left   != no else existing.left,
                        right  = right  if right  != no else existing.right,
                    )          
    # ── Внешняя рамка всей таблицы ───────────────────────────────────────
    last_data_row = base_row + len(staff) * 3 - 1
    last_data_col = 4 + total

    thick = Side(style="medium", color="000000")
    no    = Side(style=None)

    for row in range(5, last_data_row + 1):
        for col in range(1, last_data_col + 1):
            c = ws.cell(row=row, column=col)
            top    = thick if row == 5             else no
            bottom = thick if row == last_data_row else no
            left   = thick if col == 1             else no
            right  = thick if col == last_data_col else no
            if any([top, bottom, left, right]):
                existing = c.border
                c.border = Border(
                    top    = top    if top    != no else existing.top,
                    bottom = bottom if bottom != no else existing.bottom,
                    left   = left   if left   != no else existing.left,
                    right  = right  if right  != no else existing.right,
                )

    # ── Freeze panes: фиксируем 役職+氏名+メタ и строки заголовков ────────
    ws.freeze_panes = "D7"

    # Настройки печати
    ws.page_setup.orientation = "landscape"          # 横方向
    ws.page_setup.paperSize = 9                       # A4
    ws.page_setup.fitToPage = True                    # シートを1ページに印刷
    ws.page_setup.fitToWidth = 1                      # по ширине
    ws.page_setup.fitToHeight = 0                     # высота авто
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 狭い余白
    ws.page_margins = PageMargins(
        left=0.25, right=0.25,
        top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )

    # 片面印刷 (по умолчанию, но явно)
    ws.page_setup.copies = 1

    # ── Итоговый буфер ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()