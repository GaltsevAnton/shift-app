"""
Табель учёта рабочего времени.
Формат:
  Строки: сотрудники (по одной)
  Колонки: 職種 | 部署 | 氏名 | 1..31 | 出勤日数 | 総労働時間
  Ячейка дня: суммарные часы (напр. 8.5), 休 для выходных
"""

import io
import calendar
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import ReportRequest, StaffModel, DayModel

WD_JA     = ["日", "月", "火", "水", "木", "金", "土"]
FONT_NAME = "メイリオ"

C_HEADER_BG  = "1F4E79"
C_HEADER_FG  = "FFFFFF"
C_SAT_BG     = "BDD7EE"
C_SUN_BG     = "FCE4D6"
C_WEEKDAY_BG = "DDEBF7"
C_OFF_FG     = "C00000"
C_OFF_BG     = "FFE0E0"
C_ROW_ODD    = "FFFFFF"
C_ROW_EVEN   = "F2F2F2"
C_META_BG    = "F5F5F5"
C_TOTAL_BG   = "E2EFDA"   # зелёный — итоговые колонки
C_GRID       = "AAAAAA"


def _thin(color=C_GRID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=9, name=FONT_NAME):
    return Font(bold=bold, color=color, size=size, name=name)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _weekday(ym, day):
    y, m = map(int, ym.split("-"))
    return (date(y, m, day).weekday() + 1) % 7  # 0=日..6=土


def _day_data(staff: StaffModel, day: int, ym: str) -> DayModel:
    ds = f"{ym}-{day:02d}"
    for d in staff.days:
        if d.date == ds:
            return d
    return DayModel(date=ds, off=True, slots=[])


def _calc_hours(d: DayModel) -> float:
    """Суммарные часы за день по всем слотам."""
    if d.off or not d.slots:
        return 0.0
    total = 0.0
    for sl in d.slots:
        if not sl.startTime or sl.last or not sl.endTime:
            continue
        try:
            start = datetime.strptime(sl.startTime[:5], "%H:%M")
            end   = datetime.strptime(sl.endTime[:5],   "%H:%M")
            diff  = (end - start).total_seconds() / 3600
            if diff > 0:
                total += diff
        except ValueError:
            pass
    return round(total, 2)


def build(req: ReportRequest) -> bytes:
    ym    = req.ym
    y, m  = map(int, ym.split("-"))
    total = calendar.monthrange(y, m)[1]
    days  = list(range(1, total + 1))

    # Колонки итогов: total+4 = 出勤日数, total+5 = 総労働時間
    col_work_days = 4 + total
    col_total_h   = 5 + total
    last_col      = get_column_letter(col_total_h)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}年{m}月_勤怠"

    # ── Ширина колонок ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    for i in range(total):
        col = get_column_letter(4 + i)
        wd  = _weekday(ym, i + 1)
        ws.column_dimensions[col].width = 6 if wd in (0, 6) else 5.5
    ws.column_dimensions[get_column_letter(col_work_days)].width = 8
    ws.column_dimensions[get_column_letter(col_total_h)].width   = 10

    # ── Строка 1: заголовок ───────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = f"{req.hotelName}勤怠表（{y}年{m}月）"
    c.font      = _font(bold=True, size=11)
    c.alignment = _align(h="left")
    ws.row_dimensions[1].height = 20

    # ── Строка 2: заголовки ───────────────────────────────────────────────
    for col_idx, label in enumerate(["職種・役職", "部署", "氏名"], 1):
        c = ws.cell(row=2, column=col_idx)
        c.value     = label
        c.font      = _font(bold=True, color=C_HEADER_FG, size=9)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align()
        c.border    = _thin()

    for i, day in enumerate(days):
        wd     = _weekday(ym, day)
        is_sat = wd == 6
        is_sun = wd == 0
        bg     = C_SAT_BG if is_sat else (C_SUN_BG if is_sun else C_WEEKDAY_BG)
        col    = 4 + i
        c = ws.cell(row=2, column=col)
        c.value     = f"{day}\n{WD_JA[wd]}"
        c.font      = _font(
            bold=True,
            color=C_OFF_FG if is_sun else ("1F4E79" if is_sat else C_HEADER_FG),
            size=8,
        )
        c.fill      = _fill(bg)
        c.alignment = _align()
        c.border    = _thin()

    # Итоговые заголовки
    for col_idx, label in [(col_work_days, "出勤\n日数"), (col_total_h, "総労働\n時間")]:
        c = ws.cell(row=2, column=col_idx)
        c.value     = label
        c.font      = _font(bold=True, color=C_HEADER_FG, size=9)
        c.fill      = _fill("2E7D32")  # тёмно-зелёный
        c.alignment = _align()
        c.border    = _thin()

    ws.row_dimensions[2].height = 28

    # ── Данные ────────────────────────────────────────────────────────────
    for row_idx, s in enumerate(req.staff):
        r      = 3 + row_idx
        row_bg = C_ROW_ODD if row_idx % 2 == 0 else C_ROW_EVEN

        # 職種
        c = ws.cell(row=r, column=1)
        c.value = s.position or ""
        c.font = _font(size=8, color="555555")
        c.fill = _fill(C_META_BG)
        c.alignment = _align(h="left")
        c.border = _thin()

        # 部署
        c = ws.cell(row=r, column=2)
        c.value = "、".join(s.departments)
        c.font = _font(size=8, color="555555")
        c.fill = _fill(C_META_BG)
        c.alignment = _align(h="left")
        c.border = _thin()

        # 氏名
        c = ws.cell(row=r, column=3)
        c.value = s.userName
        c.font = _font(bold=True, size=9)
        c.fill = _fill(C_META_BG)
        c.alignment = _align(h="left", wrap=False)
        c.border = _thin()

        work_days = 0
        total_hours = 0.0

        for i, day in enumerate(days):
            wd     = _weekday(ym, day)
            is_sat = wd == 6
            is_sun = wd == 0
            day_bg = C_SAT_BG if is_sat else (C_SUN_BG if is_sun else row_bg)
            col    = 4 + i

            d     = _day_data(s, day, ym)
            hours = _calc_hours(d)
            c     = ws.cell(row=r, column=col)
            c.border = _thin()

            if d.off or not d.slots:
                c.value     = "休"
                c.font      = _font(bold=True, color=C_OFF_FG, size=8)
                c.fill      = _fill(C_OFF_BG)
                c.alignment = _align()
            elif hours == 0:
                # Есть слоты но нет времени (L без endTime)
                c.value     = "L"
                c.font      = _font(size=8, color="1F4E79")
                c.fill      = _fill(day_bg)
                c.alignment = _align()
                work_days  += 1
            else:
                c.value     = hours
                c.font      = _font(size=8)
                c.fill      = _fill(day_bg)
                c.alignment = _align()
                c.number_format = "0.0"
                work_days  += 1
                total_hours += hours

        # 出勤日数
        c = ws.cell(row=r, column=col_work_days)
        c.value     = work_days
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(C_TOTAL_BG)
        c.alignment = _align()
        c.border    = _thin()

        # 総労働時間
        c = ws.cell(row=r, column=col_total_h)
        c.value       = round(total_hours, 1)
        c.font        = _font(bold=True, size=9)
        c.fill        = _fill(C_TOTAL_BG)
        c.alignment   = _align()
        c.border      = _thin()
        c.number_format = "0.0"

        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "D3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()